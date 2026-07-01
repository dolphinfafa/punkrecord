"""
Finance API endpoints
"""
from typing import Optional
from urllib.parse import quote
from decimal import Decimal
from app.models.base import now_cn
from uuid import UUID
from datetime import datetime, date
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
from app.core.database import get_session
from app.core.auth import require_permission
from app.core.exceptions import AtlasException, NotFoundException
from app.core.response import success_response
from app.models.iam import User, OurEntity, OurEntityStatus
from app.models.finance import (
    FinanceAccount, FinanceTransaction, FinanceInvoice, Reimbursement,
    AccountCategory, AccountStatus, TransactionDirection, TransactionType,
    ReconcileStatus, InvoiceKind, InvoiceMedium, OCRStatus,
    ReimbursementStatus
)
from app.schemas.finance import (
    FinanceAccountCreate, FinanceAccountUpdate, FinanceAccountResponse,
    TransactionCreate, TransactionUpdate, TransactionResponse,
    InvoiceCreate, InvoiceResponse,
    ReimbursementCreate, ReimbursementResponse
)

router = APIRouter(prefix="/finance", tags=["Finance"])


# Account endpoints

@router.post("/accounts", response_model=dict)
async def create_account(
    data: FinanceAccountCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(require_permission("finance.write"))
):
    """Create finance account"""
    # Mask account number for display: show last 4 digits
    account_no_masked = None
    if data.account_no:
        if len(data.account_no) > 4:
            account_no_masked = '*' * (len(data.account_no) - 4) + data.account_no[-4:]
        else:
            account_no_masked = data.account_no

    account = FinanceAccount(
        entity_id=data.entity_id,
        account_category=AccountCategory(data.account_category),
        account_name=data.account_name,
        bank_name=data.bank_name,
        bank_branch=data.bank_branch,
        account_no_masked=account_no_masked,
        currency=data.currency,
        initial_balance=data.initial_balance,
        status=AccountStatus.ACTIVE,
        is_default=data.is_default,
        shareholder_user_id=data.shareholder_user_id
    )
    
    session.add(account)
    session.commit()
    session.refresh(account)
    
    return success_response(FinanceAccountResponse.model_validate(account))


@router.get("/accounts", response_model=dict)
async def list_accounts(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    session: Session = Depends(get_session),
    current_user: User = Depends(require_permission("finance.read"))
):
    """List finance accounts with pagination"""
    accounts = session.exec(select(FinanceAccount).where(FinanceAccount.status == AccountStatus.ACTIVE)).all()
    total = len(accounts)
    start = (page - 1) * page_size
    page_accounts = accounts[start:start + page_size]

    results = []
    for account in page_accounts:
        txns = session.exec(select(FinanceTransaction).where(FinanceTransaction.account_id == account.id)).all()

        current_balance = account.initial_balance
        for txn in txns:
            if txn.voided:
                continue
            if txn.reconcile_status not in {ReconcileStatus.COMPLETED, ReconcileStatus.RECONCILED}:
                continue
            if txn.txn_direction == TransactionDirection.IN:
                current_balance += txn.amount
            else:
                current_balance -= txn.amount

        acc_resp = FinanceAccountResponse.model_validate(account)
        acc_resp.balance = current_balance
        results.append(acc_resp)

    return success_response({
        "items": results,
        "total": total,
        "page": page,
        "page_size": page_size,
    })


@router.patch("/accounts/{account_id}", response_model=dict)
async def update_account(
    account_id: UUID,
    data: FinanceAccountUpdate,
    session: Session = Depends(get_session),
    current_user: User = Depends(require_permission("finance.write"))
):
    """Update finance account"""
    account = session.get(FinanceAccount, account_id)
    if not account:
        raise NotFoundException("未找到账户")
    
    if data.entity_id is not None:
        account.entity_id = data.entity_id
    if data.account_category is not None:
        account.account_category = AccountCategory(data.account_category)
    if data.account_name is not None:
        account.account_name = data.account_name
    if data.bank_name is not None:
        account.bank_name = data.bank_name
    if data.bank_branch is not None:
        account.bank_branch = data.bank_branch
    if data.currency is not None:
        account.currency = data.currency
    if data.initial_balance is not None:
        account.initial_balance = data.initial_balance
    if data.shareholder_user_id is not None:
        account.shareholder_user_id = data.shareholder_user_id
    if data.is_default is not None:
        account.is_default = data.is_default
    if data.status is not None:
        account.status = AccountStatus(data.status)
    if data.account_no is not None and data.account_no:
        if len(data.account_no) > 4:
            account.account_no_masked = '*' * (len(data.account_no) - 4) + data.account_no[-4:]
        else:
            account.account_no_masked = data.account_no
    elif data.account_no_masked is not None:
        account.account_no_masked = data.account_no_masked

    account.updated_at = now_cn()
    session.add(account)
    session.commit()
    session.refresh(account)
    
    return success_response(FinanceAccountResponse.model_validate(account))


# Transaction endpoints

@router.post("/transactions", response_model=dict)
async def create_transaction(
    data: TransactionCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(require_permission("finance.write"))
):
    """Create transaction"""
    # Resolve our_entity_id: if provided, validate it; otherwise, use the first active OurEntity
    our_entity_id = data.our_entity_id
    if our_entity_id:
        entity = session.get(OurEntity, our_entity_id)
        if not entity:
            # Provided ID is not a valid OurEntity (likely a counterparty ID from account.entity_id)
            our_entity_id = None
    if not our_entity_id:
        default_entity = session.exec(
            select(OurEntity).where(OurEntity.status == OurEntityStatus.ACTIVE).limit(1)
        ).first()
        if not default_entity:
            raise NotFoundException("未找到可用的主体实体，请先创建主体")
        our_entity_id = default_entity.id

    txn_type = TransactionType(data.txn_type)
    txn_direction = TransactionDirection(data.txn_direction)
    if txn_type == TransactionType.RECEIPT:
        txn_direction = TransactionDirection.IN
    elif txn_type in {TransactionType.PAYMENT, TransactionType.REIMBURSEMENT}:
        txn_direction = TransactionDirection.OUT

    transaction = FinanceTransaction(
        our_entity_id=our_entity_id,
        account_id=data.account_id,
        txn_type=txn_type,
        txn_direction=txn_direction,
        amount=data.amount,
        currency=data.currency,
        txn_date=data.txn_date,
        counterparty_id=data.counterparty_id if txn_type != TransactionType.REIMBURSEMENT else None,
        employee_user_id=data.employee_user_id if txn_type == TransactionType.REIMBURSEMENT else None,
        contract_id=data.contract_id,
        purpose=data.purpose,
        channel=data.channel,
        reference_no=data.reference_no,
        attachments=data.attachments,
        reconcile_status=ReconcileStatus(data.reconcile_status),
        created_by_user_id=current_user.id
    )
    
    session.add(transaction)
    session.commit()
    session.refresh(transaction)
    
    # Update contract pending_amount if transaction is linked to a contract
    if data.contract_id:
        from app.models.contract import Contract, ContractType
        contract = session.get(Contract, data.contract_id)
        if contract:
            # Calculate pending amount change based on contract type and transaction direction
            # Sales contract: income decreases pending (customer payment), expense increases pending (refund)
            # Purchase contract: expense decreases pending (our payment), income increases pending (supplier refund)
            
            if contract.contract_type == ContractType.SALES:
                if txn_direction == TransactionDirection.IN:
                    # Customer payment - decrease pending amount
                    contract.pending_amount -= data.amount
                else:  # 'out'
                    # Refund to customer - increase pending amount
                    contract.pending_amount += data.amount
            elif contract.contract_type == ContractType.PURCHASE:
                if txn_direction == TransactionDirection.OUT:
                    # Our payment - decrease pending amount
                    contract.pending_amount -= data.amount
                else:  # 'in'
                    # Supplier refund - increase pending amount
                    contract.pending_amount += data.amount
            # For THIRD_PARTY contracts, we don't update pending_amount for now
            
            session.add(contract)
            session.commit()
    
    return success_response(TransactionResponse.model_validate(transaction))


@router.patch("/transactions/{txn_id}", response_model=dict)
async def update_transaction(
    txn_id: UUID,
    data: TransactionUpdate,
    session: Session = Depends(get_session),
    current_user: User = Depends(require_permission("finance.write"))
):
    """Update transaction"""
    transaction = session.get(FinanceTransaction, txn_id)
    if not transaction:
        raise NotFoundException("未找到交易")

    # Track old values for contract pending_amount recalculation
    old_amount = transaction.amount
    old_direction = transaction.txn_direction
    old_contract_id = transaction.contract_id

    if data.txn_type is not None:
        txn_type = TransactionType(data.txn_type)
        transaction.txn_type = txn_type
        # Auto-set direction based on type
        if txn_type == TransactionType.RECEIPT:
            transaction.txn_direction = TransactionDirection.IN
        elif txn_type in {TransactionType.PAYMENT, TransactionType.REIMBURSEMENT}:
            transaction.txn_direction = TransactionDirection.OUT
    if data.txn_direction is not None and data.txn_type is None:
        transaction.txn_direction = TransactionDirection(data.txn_direction)
    if data.amount is not None:
        transaction.amount = data.amount
    if data.txn_date is not None:
        transaction.txn_date = data.txn_date
    if data.counterparty_id is not None:
        transaction.counterparty_id = data.counterparty_id if str(data.counterparty_id) != '' else None
    if data.employee_user_id is not None:
        transaction.employee_user_id = data.employee_user_id if str(data.employee_user_id) != '' else None
    if data.contract_id is not None:
        transaction.contract_id = data.contract_id if str(data.contract_id) != '' else None
    if data.purpose is not None:
        transaction.purpose = data.purpose
    if data.attachments is not None:
        transaction.attachments = data.attachments
    if data.reconcile_status is not None:
        transaction.reconcile_status = ReconcileStatus(data.reconcile_status)

    transaction.updated_at = now_cn()
    session.add(transaction)
    session.commit()
    session.refresh(transaction)

    # Recalculate contract pending_amount if amount/direction/contract changed
    affected_contract_ids = set()
    if old_contract_id:
        affected_contract_ids.add(old_contract_id)
    if transaction.contract_id:
        affected_contract_ids.add(transaction.contract_id)

    if affected_contract_ids and (
        data.amount is not None or data.txn_direction is not None or
        data.txn_type is not None or data.contract_id is not None
    ):
        from app.models.contract import Contract, ContractType as CType
        for cid in affected_contract_ids:
            contract = session.get(Contract, cid)
            if not contract:
                continue
            # Recalculate from scratch: pending = total - sum of relevant transactions
            txns = session.exec(
                select(FinanceTransaction).where(FinanceTransaction.contract_id == cid)
            ).all()
            paid = sum(
                t.amount if (
                    (contract.contract_type == CType.SALES and t.txn_direction == TransactionDirection.IN) or
                    (contract.contract_type == CType.PURCHASE and t.txn_direction == TransactionDirection.OUT)
                ) else -t.amount if (
                    (contract.contract_type == CType.SALES and t.txn_direction == TransactionDirection.OUT) or
                    (contract.contract_type == CType.PURCHASE and t.txn_direction == TransactionDirection.IN)
                ) else 0
                for t in txns
            )
            contract.pending_amount = contract.amount_total - paid
            session.add(contract)
        session.commit()

    return success_response(TransactionResponse.model_validate(transaction))


def _build_txn_query(
    account_id: Optional[UUID],
    txn_direction: Optional[str],
    date_from: Optional[date],
    date_to: Optional[date],
    status: Optional[str] = None,
):
    """Build shared transaction query with filters.

    status：交易列表「状态」列筛选，取值
      voided（作废）/ unreconciled / completed / reconciled。
      作废态独立；其余按对账状态且排除作废。
    """
    query = select(FinanceTransaction)
    if account_id:
        query = query.where(FinanceTransaction.account_id == account_id)
    if txn_direction:
        query = query.where(FinanceTransaction.txn_direction == txn_direction)
    if date_from:
        query = query.where(FinanceTransaction.txn_date >= date_from)
    if date_to:
        query = query.where(FinanceTransaction.txn_date <= date_to)
    if status:
        if status == "voided":
            query = query.where(FinanceTransaction.voided == True)  # noqa: E712
        else:
            query = query.where(FinanceTransaction.voided == False)  # noqa: E712
            query = query.where(FinanceTransaction.reconcile_status == status)
    return query


@router.get("/transactions", response_model=dict)
async def list_transactions(
    account_id: Optional[UUID] = Query(None),
    txn_direction: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    status: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    session: Session = Depends(get_session),
    current_user: User = Depends(require_permission("finance.read"))
):
    """List transactions"""
    query = _build_txn_query(account_id, txn_direction, date_from, date_to, status)
    query = query.order_by(FinanceTransaction.txn_date.desc())

    offset = (page - 1) * page_size
    transactions = session.exec(query.offset(offset).limit(page_size)).all()

    count_query = _build_txn_query(account_id, txn_direction, date_from, date_to, status)
    total = len(session.exec(count_query).all())

    return success_response({
        "items": [TransactionResponse.model_validate(t) for t in transactions],
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": (total + page_size - 1) // page_size
    })


@router.get("/transactions/export", response_model=None)
async def export_transactions(
    account_id: Optional[UUID] = Query(None),
    txn_direction: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    status: Optional[str] = Query(None),
    session: Session = Depends(get_session),
    current_user: User = Depends(require_permission("finance.read"))
):
    """Export transactions as Excel"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    query = _build_txn_query(account_id, txn_direction, date_from, date_to, status)
    query = query.order_by(FinanceTransaction.txn_date.desc())
    transactions = session.exec(query).all()

    # Build name maps
    from app.models.contract import Contract, Counterparty
    accounts_map = {a.id: a.account_name for a in session.exec(select(FinanceAccount)).all()}
    contracts_map = {c.id: f"{c.contract_no} - {c.name}" for c in session.exec(select(Contract)).all()}
    cp_map = {cp.id: cp.name for cp in session.exec(select(Counterparty)).all()}
    users_map = {u.id: u.display_name for u in session.exec(select(User)).all()}

    txn_type_map = {"receipt": "收款", "payment": "付款", "reimbursement": "报销"}
    status_map = {"unreconciled": "未完成", "completed": "已完成", "reconciled": "已对账"}

    wb = Workbook()
    ws = wb.active
    ws.title = "交易明细"

    headers = ["日期", "交易类型", "方向", "金额", "描述", "账户", "交易对象", "关联合同", "状态"]
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row_idx, txn in enumerate(transactions, 2):
        direction_label = "收入" if txn.txn_direction == TransactionDirection.IN else "支出"
        if txn.txn_type == TransactionType.REIMBURSEMENT:
            counterparty = users_map.get(txn.employee_user_id, "-")
        else:
            counterparty = cp_map.get(txn.counterparty_id, "-") if txn.counterparty_id else "-"

        row_data = [
            str(txn.txn_date) if txn.txn_date else "-",
            txn_type_map.get(txn.txn_type, txn.txn_type),
            direction_label,
            float(txn.amount),
            txn.purpose or "-",
            accounts_map.get(txn.account_id, "-"),
            counterparty,
            contracts_map.get(txn.contract_id, "-") if txn.contract_id else "-",
            "作废" if txn.voided else status_map.get(txn.reconcile_status, txn.reconcile_status),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 4:  # amount
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"交易明细"
    if date_from:
        filename += f"_{date_from}"
    if date_to:
        filename += f"_{date_to}"
    filename += ".xlsx"

    encoded_filename = quote(filename.encode("utf-8"))
    ascii_filename = "transactions.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f"attachment; filename={ascii_filename}; filename*=UTF-8''{encoded_filename}"
            )
        }
    )


@router.get("/transactions/{txn_id}", response_model=dict)
async def get_transaction(
    txn_id: UUID,
    session: Session = Depends(get_session),
    current_user: User = Depends(require_permission("finance.read"))
):
    """Get transaction"""
    transaction = session.get(FinanceTransaction, txn_id)
    if not transaction:
        raise NotFoundException("未找到交易")

    return success_response(TransactionResponse.model_validate(transaction))


def _contract_pending_delta(contract_type, txn_direction, amount):
    """The delta a transaction applied to contract.pending_amount when created."""
    from app.models.contract import ContractType
    if contract_type == ContractType.SALES:
        return -amount if txn_direction == TransactionDirection.IN else amount
    if contract_type == ContractType.PURCHASE:
        return -amount if txn_direction == TransactionDirection.OUT else amount
    return Decimal(0)  # third party: no pending change


def _set_transaction_voided(session: Session, transaction: FinanceTransaction, voided: bool):
    """Toggle voided and keep linked contract.pending_amount consistent."""
    if transaction.voided == voided:
        return
    transaction.voided = voided
    if transaction.contract_id:
        from app.models.contract import Contract
        contract = session.get(Contract, transaction.contract_id)
        if contract:
            delta = _contract_pending_delta(
                contract.contract_type, transaction.txn_direction, transaction.amount
            )
            # void → reverse the original effect; unvoid → re-apply it
            contract.pending_amount += -delta if voided else delta
            session.add(contract)
    session.add(transaction)
    session.commit()
    session.refresh(transaction)


@router.post("/transactions/{txn_id}/void", response_model=dict)
async def void_transaction(
    txn_id: UUID,
    session: Session = Depends(get_session),
    current_user: User = Depends(require_permission("finance.write"))
):
    """作废交易：仍展示但不计入账户余额；同步回退关联合同的待收/待付金额。"""
    transaction = session.get(FinanceTransaction, txn_id)
    if not transaction:
        raise NotFoundException("未找到交易")
    _set_transaction_voided(session, transaction, True)
    return success_response(TransactionResponse.model_validate(transaction))


@router.post("/transactions/{txn_id}/unvoid", response_model=dict)
async def unvoid_transaction(
    txn_id: UUID,
    session: Session = Depends(get_session),
    current_user: User = Depends(require_permission("finance.write"))
):
    """恢复已作废的交易。"""
    transaction = session.get(FinanceTransaction, txn_id)
    if not transaction:
        raise NotFoundException("未找到交易")
    _set_transaction_voided(session, transaction, False)
    return success_response(TransactionResponse.model_validate(transaction))


@router.delete("/transactions/{txn_id}", response_model=dict)
async def delete_voided_transaction(
    txn_id: UUID,
    session: Session = Depends(get_session),
    current_user: User = Depends(require_permission("finance.write"))
):
    """删除已作废的交易明细。正常交易必须先作废，避免误删有效流水。"""
    transaction = session.get(FinanceTransaction, txn_id)
    if not transaction:
        raise NotFoundException("未找到交易")
    if not transaction.voided:
        raise AtlasException("只能删除已作废的交易明细，请先作废后再删除")

    session.delete(transaction)
    session.commit()
    return success_response({"message": "交易明细已删除"})


# Invoice endpoints

@router.post("/invoices", response_model=dict)
async def create_invoice(
    data: InvoiceCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(require_permission("finance.write"))
):
    """Create invoice"""
    invoice = FinanceInvoice(
        our_entity_id=data.our_entity_id,
        invoice_kind=InvoiceKind(data.invoice_kind),
        invoice_medium=InvoiceMedium(data.invoice_medium),
        invoice_no=data.invoice_no,
        issue_date=data.issue_date,
        amount_with_tax=data.amount_with_tax,
        files=[],  # Files will be uploaded separately
        ocr_status=OCRStatus.PENDING,
        related_contract_id=data.related_contract_id,
        related_payment_plan_id=data.related_payment_plan_id
    )
    
    session.add(invoice)
    session.commit()
    session.refresh(invoice)
    
    return success_response(InvoiceResponse.model_validate(invoice))


@router.get("/invoices", response_model=dict)
async def list_invoices(
    invoice_kind: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    session: Session = Depends(get_session),
    current_user: User = Depends(require_permission("finance.read"))
):
    """List invoices"""
    query = select(FinanceInvoice)
    
    if invoice_kind:
        query = query.where(FinanceInvoice.invoice_kind == invoice_kind)
    
    query = query.order_by(FinanceInvoice.created_at.desc())
    
    offset = (page - 1) * page_size
    invoices = session.exec(query.offset(offset).limit(page_size)).all()
    
    count_query = select(FinanceInvoice)
    if invoice_kind:
        count_query = count_query.where(FinanceInvoice.invoice_kind == invoice_kind)
    total = len(session.exec(count_query).all())
    
    return success_response({
        "items": [InvoiceResponse.model_validate(i) for i in invoices],
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": (total + page_size - 1) // page_size
    })


# Reimbursement endpoints

@router.post("/reimbursements", response_model=dict)
async def create_reimbursement(
    data: ReimbursementCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(require_permission("finance.write"))
):
    """Create reimbursement"""
    reimbursement = Reimbursement(
        our_entity_id=data.our_entity_id,
        requester_user_id=current_user.id,
        project_id=data.project_id,
        contract_id=data.contract_id,
        total_amount=data.total_amount,
        expense_lines=data.expense_lines,
        status=ReimbursementStatus.DRAFT
    )
    
    session.add(reimbursement)
    session.commit()
    session.refresh(reimbursement)
    
    return success_response(ReimbursementResponse.model_validate(reimbursement))


@router.get("/reimbursements", response_model=dict)
async def list_reimbursements(
    status: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    session: Session = Depends(get_session),
    current_user: User = Depends(require_permission("finance.read"))
):
    """List reimbursements"""
    query = select(Reimbursement).where(Reimbursement.requester_user_id == current_user.id)
    
    if status:
        query = query.where(Reimbursement.status == status)
    
    query = query.order_by(Reimbursement.created_at.desc())
    
    offset = (page - 1) * page_size
    reimbursements = session.exec(query.offset(offset).limit(page_size)).all()
    
    count_query = select(Reimbursement).where(Reimbursement.requester_user_id == current_user.id)
    if status:
        count_query = count_query.where(Reimbursement.status == status)
    total = len(session.exec(count_query).all())
    
    return success_response({
        "items": [ReimbursementResponse.model_validate(r) for r in reimbursements],
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": (total + page_size - 1) // page_size
    })
