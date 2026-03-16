"""
Document text extraction and chunking.
Supports: PDF (PyPDF2), Word (python-docx), Excel (openpyxl), plain text,
and images (Gemini Vision).
"""
import logging
import os
from typing import List, Tuple

from app.core.config import settings

logger = logging.getLogger(__name__)


def extract_text(file_bytes: bytes, content_type: str, file_name: str) -> str:
    """Extract text content from a file based on its MIME type."""
    ct = content_type.lower()

    if ct == "application/pdf":
        return _extract_pdf(file_bytes)
    elif ct in (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    ):
        return _extract_docx(file_bytes)
    elif ct in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        return _extract_excel(file_bytes)
    elif ct.startswith("text/"):
        return file_bytes.decode("utf-8", errors="replace")
    elif ct.startswith("image/"):
        return _extract_image_via_gemini(file_bytes, content_type)
    else:
        # Try as plain text
        try:
            return file_bytes.decode("utf-8", errors="replace")
        except Exception:
            return ""


def _extract_pdf(data: bytes) -> str:
    from PyPDF2 import PdfReader
    import io

    reader = PdfReader(io.BytesIO(data))
    parts = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            parts.append(text)
    return "\n".join(parts)


def _extract_docx(data: bytes) -> str:
    from docx import Document
    import io

    doc = Document(io.BytesIO(data))
    parts = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(parts)


def _extract_excel(data: bytes) -> str:
    from openpyxl import load_workbook
    import io

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        parts.append(f"## Sheet: {sheet}")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            parts.append(" | ".join(cells))
    return "\n".join(parts)


def _extract_image_via_gemini(data: bytes, content_type: str) -> str:
    """Use Gemini Vision to extract text from images."""
    import httpx
    import base64
    import json

    api_key = settings.GEMINI_API_KEY
    if not api_key:
        logger.warning("No GEMINI_API_KEY, cannot extract image text")
        return ""

    b64 = base64.b64encode(data).decode()
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={api_key}"
    payload = {
        "contents": [{
            "parts": [
                {"text": "请提取这张图片中的所有文字内容，保持原始格式。如果图片中没有文字，请描述图片内容。"},
                {"inline_data": {"mime_type": content_type, "data": b64}},
            ]
        }]
    }

    try:
        resp = httpx.post(url, json=payload, timeout=60.0, verify=False)
        resp.raise_for_status()
        result = resp.json()
        return result["candidates"][0]["content"]["parts"][0]["text"]
    except Exception as e:
        logger.error("Gemini Vision extraction failed: %s", e)
        return ""


def chunk_text(
    text: str,
    chunk_size: int | None = None,
    chunk_overlap: int | None = None,
) -> List[Tuple[str, int]]:
    """
    Split text into overlapping chunks.

    Returns:
        List of (chunk_text, estimated_token_count) tuples.
    """
    if not text or not text.strip():
        return []

    size = chunk_size or settings.KB_CHUNK_SIZE
    overlap = chunk_overlap or settings.KB_CHUNK_OVERLAP

    chunks: List[Tuple[str, int]] = []
    start = 0
    text_len = len(text)

    while start < text_len:
        end = min(start + size, text_len)
        chunk = text[start:end]
        # Rough token estimate: ~1.5 chars per token for Chinese, ~4 for English
        token_est = max(len(chunk) // 2, 1)
        chunks.append((chunk, token_est))
        if end >= text_len:
            break
        start += size - overlap

    return chunks
